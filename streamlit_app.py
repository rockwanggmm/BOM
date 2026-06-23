import streamlit as st
import pandas as pd
import io
import openpyxl
import re
from datetime import datetime, date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="跨表大融合整合對齊工具", layout="wide")

st.title("📋 多分頁需求單「單一頁面全整合」與工單料況對齊工具")
st.markdown("""
### 🎯 最終校正版自動化邏輯（按鈕控制型）：
1. **上傳防重複跑**：必須按下下方綠色按鈕，程式才會啟動整合，避免大檔案重複讀取卡頓。
2. **全單一頁面融合**：僅在第一個分頁插入 A、B 欄並從 **P 欄 (第 16 欄)** 黏貼料況。其餘所有分頁的數據將自動垂直向下追加到第一頁尾端。
3. **日期與交期全攔截**：全面攔截「客戶交期」、「Promise Date」、「Needby Date」等欄位，將 `46021` 等數字精準還原為漂亮的 `YYYY-MM-DD` 格式。
4. **智慧相容合併欄位與表頭格式美化**：自動修正 M 欄（融合後為 O 欄）的公式關聯，並於整合完畢後將第 7、8 列的表頭樣式精準還原至與圖片一模一樣。
""")

col1, col2 = st.columns(2)
with col1:
    ecn_file = st.file_uploader("1. 請上傳 多分頁需求單檔案", type=["xlsx", "xls"])
with col2:
    status_file = st.file_uploader("2. 請上傳 工單料況檔案", type=["xlsb", "xlsx", "xls"])

# 當兩個檔案都上傳完成時，顯示執行按鈕
if ecn_file and status_file:
    st.markdown("---")
    st.info("💡 兩份檔案皆已成功讀取，請點擊下方按鈕開始進行融合對齊。")
    
    # 建立控制按鈕
    start_integration = st.button("🚀 開始全自動跨表大融合整合", type="primary", use_container_width=True)
    
    if start_integration:
        try:
            # --- 步驟 1：自動從需求單檔名擷取「需求單號」 ---
            file_name = ecn_file.name
            match_doc_no = re.search(r'([A-Za-z0-9]+)', file_name)
            doc_no = match_doc_no.group(1) if match_doc_no else "GAA260500286"

            # --- 步驟 2：讀取工單料況對照表 ---
            if status_file.name.endswith('.xlsb'):
                try:
                    raw_status_df = pd.read_excel(status_file, engine='pyxlsb', header=None)
                except Exception:
                    raw_status_df = pd.read_excel(status_file, header=None)
            else:
                raw_status_df = pd.read_excel(status_file, header=None)
                
            header_row_idx = None
            for idx in range(min(20, len(raw_status_df))):
                row_vals = [str(x).strip() for x in raw_status_df.iloc[idx].dropna().tolist()]
                if '工單號碼' in row_vals or 'Component Item' in row_vals or 'Seiban' in row_vals:
                    header_row_idx = idx
                    break
                    
            if header_row_idx is None:
                status_df = raw_status_df.copy()
                status_df.columns = [str(c).strip() for c in status_df.iloc[0]]
            else:
                status_df = raw_status_df.iloc[header_row_idx+1:].copy()
                status_df.columns = [str(c).strip() for c in raw_status_df.iloc[header_row_idx]]
                st.success(f"💡 系統提示：自動偵測到工單料況表的真正標頭位於第 **{header_row_idx + 1}** 列！")

            status_df.columns = [str(c).strip() for c in status_df.columns]
            status_cols = status_df.columns.tolist()
            
            required_status_cols = ['工單號碼', 'Component Item']
            missing_status_cols = [col for col in required_status_cols if col not in status_cols]
            
            if missing_status_cols:
                st.error(f"### ❌ 工單料況檔案格式判定失敗，依然遺失必要欄位：`{"、".join(missing_status_cols)}`")
                st.stop()
            
            # --- 步驟 3：使用 openpyxl 載入需求單主體 ---
            ecn_wb = openpyxl.load_workbook(ecn_file)
            sheet_names = ecn_wb.sheetnames
            
            st.info(f"⏳ 正在處理分頁（共 {len(sheet_names)} 個分頁），正在執行全自動單頁大融合...")
            
            def clean_upper_no(val):
                if val is None:
                    return ""
                val_str = str(val).strip()
                if val_str in ['"', '”', '同上', '']:
                    return "FOLLOW"
                if len(val_str) > 1:
                    return val_str[1:]
                return val_str

            # 🚀 加強版安全寫入函式：徹底納入「交期」與各式英文字眼
            def safe_write_cell(cell, value, col_name=""):
                if pd.isna(value) or value is None:
                    cell.value = ""
                    return
                
                # 全方位掃描可能含有日期的欄位名稱（包含客戶交期、Promise Date、Needby Date 等）
                col_name_str = str(col_name)
                is_date_col = any(k in col_name_str for k in ["日期", "交期", "Date", "Promise", "Needby", "Creation"])
                
                # 狀況 A：如果值本來就是時間/日期物件
                if isinstance(value, (datetime, date, pd.Timestamp)):
                    cell.value = value
                    cell.number_format = 'yyyy-mm-dd'
                    return
                    
                # 狀況 B：如果是日期/交期欄位，但被轉成了 46021 這樣的序列號數字
                if is_date_col and str(value).replace('.0', '').isdigit():
                    try:
                        serial_num = int(float(value))
                        parsed_date = pd.to_datetime(serial_num, unit='D', origin='1899-12-30')
                        cell.value = parsed_date
                        cell.number_format = 'yyyy-mm-dd'
                        return
                    except Exception:
                        pass
                
                # 狀況 C：其餘一般資料，正常轉字串或數值寫入
                cell.value = value

            # 鎖定第一個分頁作為整合大表
            main_sheet_name = sheet_names[0]
            main_ws = ecn_wb[main_sheet_name]
            
            # 在第一個頁面最左側插入 A, B 兩欄
            main_ws.insert_cols(1, amount=2)
            main_ws.cell(row=7, column=1, value="需求單號")
            main_ws.cell(row=7, column=2, value="工單號")
            for r in range(1, 7):
                main_ws.cell(row=r, column=1, value="")
                main_ws.cell(row=r, column=2, value="")
                
            # 工單料況表欄位標頭，固定由 P 欄 (第 16 欄) 開始寫入
            start_write_col = 16 
            for i, col_name in enumerate(status_cols):
                main_ws.cell(row=7, column=start_write_col + i, value=col_name)

            current_main_row = 9
            
            # 建立進度條優化體驗
            progress_bar = st.progress(0)
            
            # 遍歷所有的分頁進行垂直整合
            for s_idx, sheet_name in enumerate(sheet_names):
                ws = ecn_wb[sheet_name]
                
                # 更新進度條
                progress_bar.progress((s_idx + 1) / len(sheet_names))
                
                # 讀取第 6 列的原始工單文字
                work_order_text = ""
                for c in range(1, 15):
                    search_col = c + 2 if s_idx == 0 else c
                    cell_val = str(ws.cell(row=6, column=search_col).value or "")
                    if "工單號" in cell_val or "工單" in cell_val:
                        if "：" in cell_val:
                            work_order_text = cell_val.split("：")[-1].strip()
                        elif ":" in cell_val:
                            work_order_text = cell_val.split(":")[-1].strip()
                        else:
                            work_order_text = cell_val.replace("工單號", "").strip()
                        break
                
                is_multi_order = False
                if work_order_text:
                    if ',' in work_order_text or '、' in work_order_text or ' ' in work_order_text or len(re.findall(r'\([^)]*\)', work_order_text)) > 1:
                        is_multi_order = True
                else:
                    is_multi_order = True
                    
                # 破解合併欄位：十字地毯式搜索
                upper_col_idx = None
                after_col_idx = None
                
                for c in range(1, ws.max_column + 1):
                    val_r7 = str(ws.cell(row=7, column=c).value or "").strip()
                    val_r8 = str(ws.cell(row=8, column=c).value or "").strip()
                    combined_text = val_r7 + val_r8
                    
                    if '上階' in combined_text:
                        upper_col_idx = c if s_idx == 0 else c + 2
                    if '變更後' in combined_text:
                        after_col_idx = c if s_idx == 0 else c + 2

                if upper_col_idx is None: upper_col_idx = 8
                if after_col_idx is None: after_col_idx = 11

                max_r = ws.max_row
                last_valid_upper = ""
                
                for r in range(9, max_r + 1):
                    read_upper_idx = upper_col_idx if s_idx == 0 else upper_col_idx - 2
                    read_after_idx = after_col_idx if s_idx == 0 else after_col_idx - 2
                    
                    raw_upper = ws.cell(row=r, column=read_upper_idx).value
                    raw_after = ws.cell(row=r, column=read_after_idx).value
                    
                    if raw_upper is None and raw_after is None and ws.cell(row=r, column=1).value is None:
                        continue
                        
                    processed_upper = clean_upper_no(raw_upper)
                    if processed_upper == "FOLLOW":
                        use_upper = last_valid_upper
                    else:
                        use_upper = processed_upper
                        last_valid_upper = processed_upper
                    
                    if is_multi_order or not work_order_text:
                        if use_upper:
                            final_work_order = use_upper[:-3] if len(use_upper) > 3 else use_upper
                            if final_work_order.endswith('-'):
                                final_work_order = final_work_order[:-1]
                        else:
                            final_work_order = ""
                    else:
                        final_work_order = work_order_text
                    
                    if s_idx > 0:
                        for col_c in range(1, 14): 
                            orig_cell = ws.cell(row=r, column=col_c)
                            
                            # 🎯 需求 3：針對 M 欄（第 13 欄，插入 A, B 欄後對應大表第 15 欄即 O 欄）做公式攔截調整
                            cell_value_to_write = orig_cell.value
                            if col_c == 13 and isinstance(cell_value_to_write, str) and cell_value_to_write.startswith('='):
                                # 全面將公式中的 J 替換為 L (忽略大小寫)，並確保維持大寫格式符合 Excel
                                cell_value_to_write = re.sub(r'\b[jJ](\d+)\b', r'L\1', cell_value_to_write)
                                
                            target_cell = main_ws.cell(row=current_main_row, column=col_c + 2, value=cell_value_to_write)
                            if orig_cell.number_format and orig_cell.number_format != 'General':
                                target_cell.number_format = orig_cell.number_format
                    else:
                        # 處理第一個分頁時，也需要對原本的 M 欄（目前在大表是第 15 欄，即 O 欄）進行公式檢查與修改
                        orig_cell = main_ws.cell(row=r, column=15)
                        if isinstance(orig_cell.value, str) and orig_cell.value.startswith('='):
                            orig_cell.value = re.sub(r'\b[jJ](\d+)\b', r'L\1', orig_cell.value)
                    
                    main_ws.cell(row=current_main_row, column=1, value=doc_no)
                    main_ws.cell(row=current_main_row, column=2, value=final_work_order)
                    
                    val_after = str(raw_after).strip() if raw_after is not None else ""
                    matched_rows = pd.DataFrame()
                    if use_upper and val_after and val_after not in ["None", "", '"', '”', '同上']:
                        matched_rows = status_df[
                            (status_df['工單號碼'].astype(str).str.strip() == use_upper) & 
                            (status_df['Component Item'].astype(str).str.strip() == val_after)
                        ]
                    
                    if not matched_rows.empty:
                        matched_match = matched_rows.iloc[0]
                        for i, col_name in enumerate(status_cols):
                            val_to_write = matched_match[col_name]
                            target_cell = main_ws.cell(row=current_main_row, column=start_write_col + i)
                            safe_write_cell(target_cell, val_to_write, col_name=col_name)
                    
                    if s_idx == 0:
                        current_main_row = r + 1
                    else:
                        current_main_row += 1

            # 🎯 需求 1 & 2：在刪除其他分頁前，統一樣式美化整合大表的 7、8 列標頭格式
            # 建立漂亮標準的框線與字型樣式
            thin_side = Side(border_style="thin", color="D3D3D3")
            standard_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 樣式 A (需求 1 - 左圖 C欄到N欄樣式)
            left_fill = PatternFill(start_color="E6EDF5", end_color="E6EDF5", fill_type="solid") # 淡藍色調
            left_font = Font(name="Microsoft JhengHei", size=10, bold=True, color="000000")
            
            # 因插入 A、B 欄，原本的 C到N 欄在 main_ws 中變成了 第 5 欄 (E) 到 第 16 欄 (P) 之前
            for col_idx in range(5, 16):
                for row_idx in [7, 8]:
                    cell = main_ws.cell(row=row_idx, column=col_idx)
                    cell.fill = left_fill
                    cell.font = left_font
                    cell.alignment = center_align
                    cell.border = standard_border

            # 樣式 B (需求 2 - 右圖 P欄到BK欄樣式)
            right_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # 淡亮黃色調
            right_font = Font(name="Microsoft JhengHei", size=10, bold=True, color="9C640C") # 棕橘色字體
            
            # 從 P 欄 (16) 到 BK 欄 (63) 進行樣式套用
            for col_idx in range(16, 64):
                for row_idx in [7, 8]:
                    cell = main_ws.cell(row=row_idx, column=col_idx)
                    cell.fill = right_fill
                    cell.font = right_font
                    cell.alignment = center_align
                    cell.border = standard_border

            # 刪除其餘分頁
            for sheet_name in sheet_names[1:]:
                del ecn_wb[sheet_name]
                
            st.success(f"🎉 跨活頁大融合成功！已將所有分頁整合成單一工作表，且「表頭美化與M欄函式校正」已完全修復！")
            
            output = io.BytesIO()
            ecn_wb.save(output)
            processed_data = output.getvalue()
            
            st.download_button(
                label="📥 下載全自動大融合整合 Excel 檔",
                data=processed_data,
                file_name=f"{doc_no}_全自動融合整合表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        except KeyError as ke:
            st.error(f"### ❌ 比對欄位異常告警: `{ke}`")
        except Exception as e:
            st.error(f"❌ 整合過程中發生未知錯誤: {e}")
else:
    st.info("💡 提示：請同時上傳「1. 多分頁需求單」與「2. 工單料況表」以解鎖整合按鈕。")
